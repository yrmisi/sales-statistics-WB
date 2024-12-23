import mmap
from pprint import pprint
from typing import Final

from openpyxl import Workbook, load_workbook

from column_numbers_by_title import get_number_column
from income_calculation import add_the_value_of_income
from product_name import replacing_article_with_the_product_name
from sales_and_incomes import calculate_sales_turnover_and_income


def collect_sales_statistics(filename: str):
    wb: Workbook = load_workbook(filename, data_only=True)
    sheet = wb.active

    numbers_columns: dict[str, int] = get_number_column(sheet[1])
    max_row: Final[int] = sheet.max_row
    sale_str: str = "Продажа"
    return_str: str = "Возврат"
    compensation_upon_return_str: str = "Добровольная компенсация при возврате"
    total_sales: float = 0
    total_sales_after_deduction: float = 0
    total_logistics: float = 0
    total_return: float = 0
    total_return_after_deduction: float = 0
    total_fines: float = 0
    total_storage: float = 0
    total_compensation_upon_return: float = 0
    cost_of_paid_acceptance: float = 0
    sales_statistics: dict = {}

    for row in range(2, max_row + 1):
        delivery_number: int = sheet.cell(
            row=row, column=numbers_columns["Номер поставки"]
        ).value
        article: str = sheet.cell(
            row=row, column=numbers_columns["Артикул поставщика"]
        ).value
        reasons_for_payment: str = sheet.cell(
            row=row, column=numbers_columns["Обоснование для оплаты"]
        ).value
        price: float = sheet.cell(
            row=row, column=numbers_columns["Вайлдберриз реализовал Товар (Пр)"]
        ).value
        price_after_deduction: float = sheet.cell(
            row=row,
            column=numbers_columns["К перечислению Продавцу за реализованный Товар"],
        ).value
        logistics: float = sheet.cell(
            row=row - 1,
            column=numbers_columns["Услуги по доставке товара покупателю"],
        ).value
        quantity: int = sheet.cell(row=row, column=numbers_columns["Кол-во"]).value

        if row != 2:
            total_logistics += logistics

        if reasons_for_payment == sale_str:
            total_sales += price
            total_sales_after_deduction += price_after_deduction

            if delivery_number not in sales_statistics:
                sales_statistics.setdefault(
                    delivery_number,
                    {article: [price, price_after_deduction, logistics, quantity]},
                )
            elif article not in sales_statistics[delivery_number]:
                sales_statistics[delivery_number][article]: list[float | int] = [
                    price,
                    price_after_deduction,
                    logistics,
                    quantity,
                ]
            else:
                sum_price: float = sales_statistics[delivery_number][article][0] + price
                sum_price_after_deduction: float = (
                    sales_statistics[delivery_number][article][1]
                    + price_after_deduction
                )
                sum_logistics: float = (
                    sales_statistics[delivery_number][article][2] + logistics
                )
                sum_quantity: int = (
                    sales_statistics[delivery_number][article][3] + quantity
                )
                sales_statistics[delivery_number][article]: list[float | int] = [
                    round(sum_price, 2),
                    round(sum_price_after_deduction, 2),
                    round(sum_logistics, 2),
                    sum_quantity,
                ]
        elif reasons_for_payment == return_str:
            total_return += price
            total_return_after_deduction += price_after_deduction

        elif reasons_for_payment == compensation_upon_return_str:
            total_compensation_upon_return += price_after_deduction

        fines: float = sheet.cell(
            row=row, column=numbers_columns["Общая сумма штрафов"]
        ).value
        total_fines += fines
        storage: float = sheet.cell(row=row, column=numbers_columns["Хранение"]).value
        total_storage += storage
        paid_acceptance: float = sheet.cell(
            row=row, column=numbers_columns["Платная приемка"]
        ).value
        cost_of_paid_acceptance += paid_acceptance

    sales_statistics: dict[int, dict[str, list[float | int]]] = add_the_value_of_income(
        sales_statistics
    )
    sales_statistics: dict[int, dict[str, list[float | int]]] = (
        replacing_article_with_the_product_name(sales_statistics)
    )
    print(", ".join(filename.split("/")[5:]).replace(".xlsx", "").capitalize())

    total_sales: float = round(total_sales - total_return, 2)
    total_sales_after_deduction: float = round(
        total_sales_after_deduction
        + total_compensation_upon_return
        - total_return_after_deduction,
        2,
    )
    total_logistics: float = round(total_logistics, 2)
    total_fines: float = round(total_fines, 2)
    total_storage: float = round(total_storage, 2)
    total_payable: float = round(
        total_sales_after_deduction
        - total_logistics
        - total_fines
        - total_storage
        - cost_of_paid_acceptance,
        2,
    )
    percentage_of_expenses: float = round(total_payable * 100 / total_sales, 2)
    print(
        f"{total_sales=}, {total_return=}, {total_sales_after_deduction=}, {total_return_after_deduction=}, \n"
        f"{total_logistics=}, {total_fines=} {total_storage=}, {cost_of_paid_acceptance=}, \n"
        f"{total_payable=}, {percentage_of_expenses=}%"
    )
    print(
        "price, price after deduction, logistics, quantity, мargin, total margin, total cost price"
    )
    pprint(sales_statistics, width=120)

    sale_of_goods, total_incomes, total_cost_price = (
        calculate_sales_turnover_and_income(sales_statistics)
    )
    print(f"{sale_of_goods=}, {total_incomes=}, {total_cost_price=} \n")
    return (
        sales_statistics,
        sale_of_goods,
        total_incomes,
        total_cost_price,
        total_sales,
        total_sales_after_deduction,
        total_payable,
    )
